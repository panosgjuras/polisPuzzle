"""Download settlement demographic matrices from the ELSTAT 2021 census."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from io import BytesIO
from itertools import product
import time
import unicodedata
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import pandas as pd
import numpy as np
from openpyxl import load_workbook


ELSTAT_URL_TEMPLATE = (
    "https://www.statistics.gr/el/statistics"
    "?p_p_id=documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ"
    "&p_p_lifecycle=2&p_p_state=normal&p_p_mode=view"
    "&p_p_cacheability=cacheLevelPage"
    "&p_p_col_id=column-2&p_p_col_count=4&p_p_col_pos=2"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_javax.faces.resource=document"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_ln=downloadResources"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_documentID={document_id}"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_locale=el"
)

GENDER_AGE_URL = ELSTAT_URL_TEMPLATE.format(document_id="568507")  # table G01
GENDER_EDUCATION_URL = ELSTAT_URL_TEMPLATE.format(document_id="568504")  # table G04
AGE_EDUCATION_URL = ELSTAT_URL_TEMPLATE.format(document_id="568511")  # table B02
EDUCATION_EMPLOYMENT_URL = ELSTAT_URL_TEMPLATE.format(document_id="532737")  # B03
GENDER_EMPLOYMENT_URL = ELSTAT_URL_TEMPLATE.format(document_id="532738")  # B02
AGE_EMPLOYMENT_URL = ELSTAT_URL_TEMPLATE.format(document_id="532739")  # B01
CAR_COUNT_URL = ELSTAT_URL_TEMPLATE.format(document_id="532688")  # G13

SETTLEMENT_LEVEL = 8
MUNICIPALITY_LEVEL = 5
MUNICIPAL_UNIT_LEVEL = 6
COMMUNITY_LEVEL = 7
AGE_GROUPS = (
    "0-4", "5-9", "10-14", "15-19", "20-24", "25-29", "30-34", "35-39",
    "40-44", "45-49", "50-54", "55-59", "60-64", "65-69", "70-74", "75+",
)
COARSE_AGE_GROUPS = ("0-14", "15-29", "30-44", "45-59", "60-74", "75+")
EDUCATION_LEVELS = (
    "Διδακτορικό/Μεταπτυχιακό/Πανεπιστήμιο/ΑΤΕΙ",
    "Μεταδευτεροβάθμια (ΙΕΚ, Κολέγια)",
    "Απολυτήριο Λυκείου",
    "Πτυχίο Επαγγελματικής Σχολής / Γυμνάσιο",
    "Απολυτήριο Δημοτικού",
    "Χωρίς απολυτήριο Δημοτικού (γνώση γραφής-ανάγνωσης)",
    "Μη κατατασσόμενοι (γεννηθέντες μετά 1/1/2016)",
)
EDUCATION_LEVELS_FULL = (
    "Διδακτορικό / Μεταπτυχιακό / Πτυχίο Πανεπιστημίου-Πολυτεχνείου, ΑΤΕΙ, "
    "ΑΣΠΑΙΤΕ, Ανώτερων Επαγγελματικών Σχολών και ισότιμων σχολών",
    "Πτυχίο μεταδευτεροβάθμιας εκπαίδευσης (ΙΕΚ, Κολέγια κλπ.)",
    "Απολυτήριο Λυκείου (Γενικού, Επαγγελματικού, Εκκλησιαστικού κ.λπ. ή "
    "εξαταξίου Γυμνασίου)",
    "Πτυχίο Επαγγελματικών Σχολών / Απολυτήριο τριτάξιου Γυμνασίου",
    "Απολυτήριο Δημοτικού",
    "Εγκατέλειψε ή δε φοίτησε στο  Δημοτικό, αλλά γνωρίζει γραφή και ανάγνωση / "
    "Ολοκλήρωσε την προσχολική αγωγή / Δεν γνωρίζει γραφή και ανάγνωση",
    "Μη κατατασσόμενοι (άτομα γεννηθέντα μετά την 1/1/2016)",
)
EMPLOYMENT_LEVELS = (
    "Απασχολούμενοι",
    "Άνεργοι",
    "Μαθητές-σπουδαστές",
    "Συνταξιούχοι",
    "Λοιποί",
)


def _download_xlsx(url: str, *, timeout: int = 120, attempts: int = 4) -> BytesIO:
    """Download an XLSX URL into memory, retrying transient ELSTAT failures."""
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; PolisPuzzle/0.1)",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "Referer": "https://www.statistics.gr/el/statistics",
            "Connection": "close",
        },
    )
    error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            with urlopen(request, timeout=timeout) as response:
                data = response.read()
            if not data.startswith(b"PK"):
                raise ValueError("ELSTAT did not return a valid XLSX file")
            return BytesIO(data)
        except (HTTPError, URLError, ConnectionError, TimeoutError, ValueError) as exc:
            error = exc
            if attempt < attempts:
                time.sleep(2 * attempt)
    raise RuntimeError(f"Could not download ELSTAT workbook from {url}") from error


def _rows(workbook: BytesIO, sheet_name: str, min_row: int = 6):
    workbook.seek(0)
    sheet = load_workbook(workbook, read_only=True, data_only=True)[sheet_name]
    yield from sheet.iter_rows(min_row=min_row, values_only=True)


def _geographic_rows(url, sheet, level, code, *, min_row=6):
    """Return ``label: row`` for one ELSTAT geographic unit."""
    return {
        str(row[2]).strip(): row
        for row in _rows(_download_xlsx(url), sheet, min_row)
        if row[0] == level and _code(row[1]) == str(code)
    }


def _text_key(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value).strip().lower().replace("ς", "σ"))
    return " ".join("".join(c for c in text if unicodedata.category(c) != "Mn").split())


def _code(value: object) -> str:
    """Preserve ELSTAT geographic identifiers read as either numbers or text."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_name(value: object) -> str:
    return str(value).split(",", 1)[0].strip()


def build_probability_matrix(row, blocks, categories, dimensions):
    """Build and normalize a two-dimensional ELSTAT probability matrix.

    Parameters
    ----------
    row : sequence
        One ELSTAT worksheet row.
    blocks : sequence of (label, start_column)
        Row labels and the zero-based start of each contiguous value block.
    categories : sequence
        Output column labels.
    dimensions : (str, str)
        Names of the row and column dimensions.
    """
    matrix = pd.DataFrame(
        {label: [row[start + i] or 0 for i in range(len(categories))]
         for label, start in blocks},
        index=pd.Index(categories, name=dimensions[1]),
    ).T
    matrix.index.name = dimensions[0]
    return matrix / matrix.to_numpy().sum()


def reconcile_totals(*margins, reference=0):
    """Scale margins to a shared total and return ``(scaled, factors)``.

    Each input is converted to float without changing its labels. The margin at
    ``reference`` supplies the common total; all other margins are scaled
    proportionally. This isolates ELSTAT cross-table total reconciliation from
    the IPF algorithm.
    """
    margins = tuple(margin.astype(float) for margin in margins)
    totals = np.array([margin.sum() for margin in margins], dtype=float)
    if (totals <= 0).any():
        raise ValueError("All margins must have positive totals")
    factors = totals[reference] / totals
    return tuple(margin * factor for margin, factor in zip(margins, factors)), factors


def list_elstat_settlements(*, gender_age_url: str = GENDER_AGE_URL) -> pd.DataFrame:
    """List settlements and their parent administrative areas.

    Parameters
    ----------
    gender_age_url : str
        ELSTAT G01 workbook URL.
    Returns
    -------
    pandas.DataFrame
        Settlement labels/codes plus municipality, municipal-unit and community
        names/codes.
    """
    book = _download_xlsx(gender_age_url)
    rows = [row for row in _rows(book, "Γ01") if row[0] is not None and row[1] is not None]
    names_by_level_and_code = {
        (row[0], _code(row[1])): str(row[2]).strip()
        for row in rows
        if row[0] in {MUNICIPALITY_LEVEL, MUNICIPAL_UNIT_LEVEL, COMMUNITY_LEVEL}
    }

    records = []
    for row in rows:
        if row[0] != SETTLEMENT_LEVEL:
            continue
        settlement_code = _code(row[1])
        records.append(
            {
                "settlement_code": settlement_code,
                "settlement_name": _clean_name(row[2]),
                "settlement_label": str(row[2]).strip(),
                "municipality_code": settlement_code[:7],
                "municipality_name": names_by_level_and_code.get(
                    (MUNICIPALITY_LEVEL, settlement_code[:7])
                ),
                "municipal_unit_code": settlement_code[:9],
                "municipal_unit_name": names_by_level_and_code.get(
                    (MUNICIPAL_UNIT_LEVEL, settlement_code[:9])
                ),
                "community_code": settlement_code[:11],
                "community_name": names_by_level_and_code.get(
                    (COMMUNITY_LEVEL, settlement_code[:11])
                ),
            }
        )
    return pd.DataFrame(records).drop_duplicates("settlement_code").reset_index(drop=True)


def _find_settlement(rows: list[tuple], settlement: str) -> tuple:
    requested = str(settlement).strip()
    by_code = [row for row in rows if _code(row[1]) == requested]
    if by_code:
        return by_code[0]

    key = _text_key(requested)
    by_name = [row for row in rows if _text_key(_clean_name(row[2])) == key]
    if not by_name:
        raise LookupError(f"Settlement {settlement!r} was not found")
    if len(by_name) > 1:
        choices = ", ".join(f"{_clean_name(r[2])} ({_code(r[1])})" for r in by_name)
        raise LookupError(f"Settlement name {settlement!r} is ambiguous; use a code: {choices}")
    return by_name[0]


def get_settlement_demographics(
    settlement: str,
    *,
    gender_age_url: str = GENDER_AGE_URL,
    gender_education_url: str = GENDER_EDUCATION_URL,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Download the two observed settlement demographic matrices.

    Parameters
    ----------
    settlement : str
        Exact ELSTAT settlement code or accent-insensitive exact name.
    gender_age_url, gender_education_url : str
        ELSTAT G01 and G04 workbook URLs.
    Returns
    -------
    (pandas.DataFrame, pandas.DataFrame)
        Gender-age and gender-education probability matrices. Both sum to one
        and carry settlement metadata in ``DataFrame.attrs``.
    """
    gender_age_book = _download_xlsx(gender_age_url)
    g01_rows = [
        row for row in _rows(gender_age_book, "Γ01")
        if row[0] == SETTLEMENT_LEVEL
    ]
    settlement_row = _find_settlement(g01_rows, settlement)
    settlement_code = _code(settlement_row[1])
    municipality_code = settlement_code[:7]

    gender_age = build_probability_matrix(
        settlement_row,
        (("Άρρενες", 21), ("Θήλεις", 38)),
        AGE_GROUPS,
        ("gender", "age_group"),
    )

    gender_education_book = _download_xlsx(gender_education_url)
    g04_rows = [
        row for row in _rows(gender_education_book, "Γ04")
        if row[0] == SETTLEMENT_LEVEL
    ]
    education_row = next(
        (row for row in g04_rows if _code(row[1]) == settlement_code),
        None,
    )
    gender_education = build_probability_matrix(
        education_row,
        (("Άρρενες", 12), ("Θήλεις", 20)),
        EDUCATION_LEVELS,
        ("gender", "education"),
    )

    metadata = {
        "settlement_code": settlement_code,
        "settlement_name": _clean_name(settlement_row[2]),
        "municipality_code": municipality_code,
        "settlement_population": int(settlement_row[3]),
    }
    gender_age.attrs.update(metadata)
    gender_education.attrs.update(metadata)
    gender_age.attrs["unit"] = gender_education.attrs["unit"] = "proportion"
    return gender_age, gender_education


def _municipality_age_education_seed(
    municipality_code: str,
    age_education_url: str,
) -> pd.DataFrame:
    """Read a municipality's observed B02 joint distribution for use as a seed."""
    municipality_rows = _geographic_rows(
        age_education_url, "Β02", MUNICIPALITY_LEVEL, municipality_code
    )
    missing = set(COARSE_AGE_GROUPS) - municipality_rows.keys()
    if missing:
        raise LookupError(
            f"Municipality {municipality_code} is missing age groups in ELSTAT B02: "
            f"{sorted(missing)}"
        )
    return pd.DataFrame(
        [[municipality_rows[age][4 + i] or 0 for i in range(len(EDUCATION_LEVELS))]
         for age in COARSE_AGE_GROUPS],
        index=pd.Index(COARSE_AGE_GROUPS, name="age_group"),
        columns=pd.Index(EDUCATION_LEVELS, name="education"), dtype=float,
    )


def get_ipf_seed(dimensions, metadata, *, age_to_coarse=None):
    """Download the ELSTAT seed associated with a pair of dimensions.

    Supported pairs are age-education, age-employment, gender-employment and
    education-employment. ``metadata`` must contain ``municipality_code``.
    Detailed age rows are created from municipality coarse rows using the
    ``coarse_age_groups`` mapping from the socioeconomic YAML configuration.
    """
    dimensions = tuple(dimensions)
    municipality = str(metadata["municipality_code"])
    pair = frozenset(dimensions)
    employment_columns = (5, 6, 8, 9, 10)

    if pair == {"age_group", "education"}:
        seed = _municipality_age_education_seed(municipality, AGE_EDUCATION_URL)
    elif pair == {"gender", "employment"}:
        rows = _geographic_rows(
            GENDER_EMPLOYMENT_URL, "Β02", MUNICIPALITY_LEVEL, municipality, min_row=7
        )
        seed = pd.DataFrame(
            [[rows[label][column] or 0 for column in employment_columns]
             for label in ("Άρρενες", "Θήλεις")],
            index=pd.Index(("Άρρενες", "Θήλεις"), name="gender"),
            columns=pd.Index(EMPLOYMENT_LEVELS, name="employment"), dtype=float,
        )
    elif pair == {"education", "employment"}:
        rows = _geographic_rows(
            EDUCATION_EMPLOYMENT_URL, "Β03", MUNICIPALITY_LEVEL, municipality,
            min_row=7,
        )
        seed = pd.DataFrame(
            [[rows[label][column] or 0 for column in employment_columns]
             for label in EDUCATION_LEVELS_FULL],
            index=pd.Index(EDUCATION_LEVELS, name="education"),
            columns=pd.Index(EMPLOYMENT_LEVELS, name="employment"), dtype=float,
        )
    elif pair == {"age_group", "employment"}:
        rows = _geographic_rows(
            AGE_EMPLOYMENT_URL, "Β01", MUNICIPALITY_LEVEL, municipality, min_row=8
        )
        values = lambda row: [
            row[5] or 0, row[6] or 0, row[10] or 0, row[11] or 0,
            sum((row[column] or 0) for column in (12, 13, 14)),
        ]
        seed = pd.DataFrame(
            [values(rows[age]) for age in COARSE_AGE_GROUPS],
            index=pd.Index(COARSE_AGE_GROUPS, name="age_group"),
            columns=pd.Index(EMPLOYMENT_LEVELS, name="employment"), dtype=float,
        )
    else:
        raise ValueError(f"No automatic ELSTAT seed for dimensions {dimensions}")

    if "age_group" in dimensions:
        if not age_to_coarse or set(AGE_GROUPS) - set(age_to_coarse):
            raise ValueError("Provide the complete coarse_age_groups YAML mapping")
        seed = pd.DataFrame(
            [seed.loc[age_to_coarse[age]].to_numpy() for age in AGE_GROUPS],
            index=pd.Index(AGE_GROUPS, name="age_group"), columns=seed.columns,
        )
    return seed if tuple(seed.axes[i].name for i in (0, 1)) == dimensions else seed.T


def _prepare_ipf_inputs(
    seed: np.ndarray,
    targets: Mapping[tuple[str, ...], np.ndarray],
    dimensions: Sequence[str],
):
    """Validate and arrange IPF constraints in NumPy axis order."""
    fitted = np.asarray(seed, dtype=float).copy()
    dimensions = tuple(dimensions)
    if fitted.ndim != len(dimensions) or len(set(dimensions)) != len(dimensions):
        raise ValueError("seed dimensions and unique dimension names must match")
    if not dimensions or not targets:
        raise ValueError("IPF requires dimensions and at least one target margin")
    if not np.isfinite(fitted).all() or (fitted < 0).any() or fitted.sum() <= 0:
        raise ValueError("seed must contain finite, non-negative values with a positive sum")

    axis = {name: i for i, name in enumerate(dimensions)}
    prepared, totals = [], []
    for names, values in targets.items():
        names = (names,) if isinstance(names, str) else tuple(names)
        if not names or len(set(names)) != len(names) or set(names) - set(dimensions):
            raise ValueError(f"Invalid target dimensions: {names}")
        values = np.asarray(values, dtype=float)
        expected = tuple(fitted.shape[axis[name]] for name in names)
        if values.shape != expected:
            raise ValueError(f"Target {names} has shape {values.shape}; expected {expected}")
        if not np.isfinite(values).all() or (values < 0).any() or values.sum() <= 0:
            raise ValueError(f"Target {names} must be finite, non-negative and positive")

        order = np.argsort([axis[name] for name in names])
        names = tuple(names[i] for i in order)
        values = values.transpose(order) if values.ndim > 1 else values
        axes = tuple(axis[name] for name in names)
        current = fitted.sum(axis=tuple(i for i in range(fitted.ndim) if i not in axes))
        if ((current == 0) & (values > 0)).any():
            raise ValueError(f"Seed structural zeros make target {names} infeasible")
        prepared.append((names, axes, values))
        totals.append(float(values.sum()))

    if not np.allclose(totals, totals[0], rtol=1e-9, atol=1e-12):
        raise ValueError(f"All target margins must have the same total; got {totals}")
    return fitted, dimensions, prepared


def validate_ipf_inputs(
    seed: np.ndarray,
    targets: Mapping[tuple[str, ...], np.ndarray],
    dimensions: Sequence[str],
) -> None:
    """Validate an arbitrary-dimensional IPF problem without fitting it.

    Checks dimension uniqueness, seed/target shapes, finite non-negative values,
    common target totals, and immediate structural-zero infeasibility. Raises
    ``ValueError`` when a condition fails and returns ``None`` otherwise.
    """
    _prepare_ipf_inputs(seed, targets, dimensions)


def _margin(array, axes):
    """Sum an array over every axis except ``axes``."""
    return array.sum(axis=tuple(set(range(array.ndim)) - set(axes)))


def _expand(array, axes, ndim):
    """Reshape a margin so it broadcasts onto its source array."""
    shape = [1] * ndim
    for size, axis in zip(array.shape, axes):
        shape[axis] = size
    return array.reshape(shape)


def run_ipf(
    seed: np.ndarray,
    targets: Mapping[tuple[str, ...], np.ndarray],
    dimensions: Sequence[str],
    *,
    max_iterations: int = 1_000,
    tolerance: float = 1e-8,
    return_diagnostics: bool = False,
):
    """Run the numerical IPF algorithm on an array of any dimensionality.

    Target keys name the retained dimensions. For example, with dimensions
    ``("gender", "age", "education")``, a ``("gender", "age")`` target must
    have shape ``(n_gender, n_age)``. Targets may be one- or multidimensional.
    Returns the fitted array, or ``(array, diagnostics)`` when
    ``return_diagnostics=True``.
    """
    fitted, dimensions, constraints = _prepare_ipf_inputs(seed, targets, dimensions)
    for iteration in range(1, max_iterations + 1):
        for _, axes, target in constraints:
            current = _margin(fitted, axes)
            ratio = np.divide(target, current, out=np.zeros_like(target), where=current != 0)
            fitted *= _expand(ratio, axes, fitted.ndim)

        errors = [
            float(np.max(np.abs(_margin(fitted, axes) - target)))
            for _, axes, target in constraints
        ]
        if max(errors) <= tolerance:
            diagnostics = {
                "iterations": iteration,
                "maximum_error": max(errors),
                "tolerance": tolerance,
                "dimensions": dimensions,
            }
            return (fitted, diagnostics) if return_diagnostics else fitted

    raise RuntimeError(f"IPF did not converge after {max_iterations} iterations")


def _association_seed(dimensions, targets, pairwise, max_iterations, tolerance):
    """Combine pairwise association factors into one multidimensional seed."""
    axes = {name: i for i, name in enumerate(dimensions)}
    seed = np.ones(tuple(len(targets[name]) for name in dimensions))
    for name in dimensions:
        seed *= _expand(targets[name].to_numpy(), (axes[name],), seed.ndim)

    for pair, matrix in pairwise.items():
        pair = tuple(pair)
        fitted = run_ipf(
            matrix.reindex(index=targets[pair[0]].index,
                           columns=targets[pair[1]].index).to_numpy(float),
            {(name,): targets[name].to_numpy() for name in pair},
            pair,
            max_iterations=max_iterations,
            tolerance=tolerance,
        )
        expected = np.outer(*(targets[name] for name in pair))
        seed *= _expand(
            fitted / expected, tuple(axes[name] for name in pair), seed.ndim
        )
    return seed


def ipf(
    dimensions,
    targets,
    *,
    metadata=None,
    mapping=None,
    seed=None,
    pairwise=None,
    constraints="double",
    max_iterations=1_000,
    tolerance=1e-8,
):
    """Run the project IPF workflow for any requested dimensions.

    Parameters
    ----------
    dimensions : sequence of str
        Output dimension order, such as ``("age_group", "education")``.
    targets : mapping of str to pandas.Series
        Known one-dimensional margins. Missing two-dimensional margins are
        derived from the downloaded seed.
    metadata : mapping, optional
        Settlement metadata containing ``municipality_code``. Required when an
        ELSTAT seed is selected automatically.
    mapping : mapping, optional
        Socioeconomic mapping. It standardizes automatic seeds and supplies
        ``coarse_age_groups``.
    seed : pandas.DataFrame or ndarray, optional
        Explicit seed. For supported two-dimensional pairs it is downloaded
        automatically when omitted.
    pairwise : mapping, optional
        Pair-dimension tuples to standardized DataFrames. Required to construct
        a multidimensional association seed when ``seed`` is omitted.
    constraints : {"single", "double", "all"}
        Constrain the first, first two, or every dimension respectively.

    Returns
    -------
    pandas.DataFrame
        A matrix for two dimensions, otherwise a long-form probability table.
    """
    dimensions = tuple(dimensions)
    metadata = dict(metadata or {})
    targets = {name: values.astype(float) / values.sum() for name, values in targets.items()}

    if seed is None and len(dimensions) == 2:
        seed = get_ipf_seed(
            dimensions, metadata,
            age_to_coarse=(mapping or {}).get("coarse_age_groups"),
        )
        if mapping:
            from .standardize import standardize_matrix
            seed = standardize_matrix(seed, *dimensions, mapping)
    elif seed is None:
        if not pairwise or set(dimensions) - set(targets):
            raise ValueError("Multidimensional IPF requires all targets and pairwise matrices")
        seed = _association_seed(
            dimensions, targets, pairwise, max_iterations, tolerance
        )

    if isinstance(seed, pd.DataFrame):
        for axis, name in enumerate(dimensions):
            if name not in targets:
                values = seed.sum(axis=1 if axis == 0 else 0)
                targets[name] = values / values.sum()
        seed = seed.reindex(index=targets[dimensions[0]].index,
                            columns=targets[dimensions[1]].index)
        labels = {dimensions[0]: seed.index, dimensions[1]: seed.columns}
        values = seed.to_numpy(float)
    else:
        values = np.asarray(seed, dtype=float)
        labels = {
            name: (targets[name].index if name in targets
                   else pd.RangeIndex(values.shape[axis], name=name))
            for axis, name in enumerate(dimensions)
        }

    count = {"single": 1, "double": 2, "all": len(dimensions)}.get(constraints)
    if count is None:
        raise ValueError("constraints must be 'single', 'double' or 'all'")
    constrained = dimensions[:count]
    scaled, factors = reconcile_totals(*(targets[name] for name in constrained))
    targets.update(zip(constrained, scaled))
    values, diagnostics = run_ipf(
        values,
        {(name,): targets[name].to_numpy() for name in constrained},
        dimensions,
        max_iterations=max_iterations,
        tolerance=tolerance,
        return_diagnostics=True,
    )
    values /= values.sum()

    if len(dimensions) == 2:
        result = pd.DataFrame(values, index=labels[dimensions[0]],
                              columns=labels[dimensions[1]])
        result.index.name, result.columns.name = dimensions
    else:
        result = pd.DataFrame(product(*(labels[name] for name in dimensions)),
                              columns=dimensions)
        result["probability"] = values.reshape(-1)
    result.attrs.update(metadata, unit="proportion", method="iterative proportional fitting",
                        constraint=constraints, reconciliation_factors=factors.tolist(),
                        **diagnostics)
    return result


def generate_agents_from_joint_distribution(
    joint_distribution: pd.DataFrame,
    population_percentage: float,
    *,
    population: int | None = None,
    random_seed: int | None = None,
) -> pd.DataFrame:
    """Sample standardized agents from a joint probability table.

    ``population_percentage`` is on the 0--100 scale. Population is read from
    table metadata unless supplied explicitly. Sampling is reproducible with
    ``random_seed``. The returned DataFrame is indexed by ``pid`` and contains
    city, gender, age group, education and employment.
    """
    required = ["gender", "age_group", "education", "employment", "probability"]
    missing = set(required) - set(joint_distribution.columns)
    if missing:
        raise ValueError(f"joint_distribution is missing columns: {sorted(missing)}")
    if not 0 <= population_percentage <= 100:
        raise ValueError("population_percentage must be between 0 and 100")

    if population is None:
        population = joint_distribution.attrs.get("settlement_population")
    if population is None:
        raise ValueError("Provide population because the joint table has no population metadata")
    population = int(round(float(population)))
    if population < 0:
        raise ValueError("population must be non-negative")
    number_of_agents = int(round(population * population_percentage / 100))

    probabilities = joint_distribution["probability"].to_numpy(dtype=float)
    if np.isnan(probabilities).any() or (probabilities < 0).any() or probabilities.sum() <= 0:
        raise ValueError("joint_distribution contains invalid probabilities")
    probabilities = probabilities / probabilities.sum()

    rng = np.random.default_rng(random_seed)
    sampled_positions = rng.choice(
        len(joint_distribution),
        size=number_of_agents,
        replace=True,
        p=probabilities,
    )
    agents = joint_distribution.iloc[sampled_positions][required[:-1]].reset_index(drop=True)
    agents.insert(
        0,
        "city",
        joint_distribution.attrs.get("settlement_name", "unknown"),
    )
    agents.index = pd.RangeIndex(1, number_of_agents + 1, name="pid")
    agents.attrs.update(
        settlement_code=joint_distribution.attrs.get("settlement_code"),
        settlement_population=population,
        population_percentage=population_percentage,
        random_seed=random_seed,
    )
    return agents


def assign_car_count_to_agents(
    agents: pd.DataFrame,
    *,
    settlement_code: str | None = None,
    car_count_url: str = CAR_COUNT_URL,
    random_seed: int | None = None,
) -> pd.DataFrame:
    """Add raw municipal-unit household car categories to agents.

    The municipal unit is derived from the settlement code. G13 household
    probabilities are transferred to the settlement and sampled independently
    per agent. Values remain raw so ``add_car_count`` can standardize them later.
    This is a person-level approximation to household-level source data.
    """
    if not isinstance(agents, pd.DataFrame):
        raise TypeError("agents must be a pandas DataFrame")
    if settlement_code is None:
        settlement_code = agents.attrs.get("settlement_code")
    if settlement_code is None:
        raise ValueError(
            "Provide settlement_code because agents have no settlement metadata"
        )
    settlement_code = str(settlement_code).strip()
    if len(settlement_code) < 9:
        raise ValueError("settlement_code must contain at least nine digits")
    municipal_unit_code = settlement_code[:9]

    rows = _geographic_rows(
        car_count_url, "Γ13", MUNICIPAL_UNIT_LEVEL, municipal_unit_code, min_row=7
    )
    municipal_unit_row = next(iter(rows.values()), None)
    if municipal_unit_row is None:
        raise LookupError(
            f"Municipal unit {municipal_unit_code} was not found in ELSTAT G13"
        )

    categories = np.array(
        ("Χωρίς αυτοκίνητο", "1 αυτοκίνητο", "2 αυτοκίνητα", "3+ αυτοκίνητα"),
        dtype=object,
    )
    household_counts = np.array(
        [municipal_unit_row[column] or 0 for column in (4, 6, 7, 8)],
        dtype=float,
    )
    if (household_counts < 0).any() or household_counts.sum() <= 0:
        raise ValueError(
            f"Municipal unit {municipal_unit_code} has invalid G13 household counts"
        )
    probabilities = household_counts / household_counts.sum()

    result = agents.copy()
    rng = np.random.default_rng(random_seed)
    result["car_count"] = rng.choice(
        categories,
        size=len(result),
        replace=True,
        p=probabilities,
    )
    result.attrs.update(
        municipal_unit_code=municipal_unit_code,
        municipal_unit_name=str(municipal_unit_row[2]).strip(),
        car_count_source="ELSTAT G13 household distribution",
        car_count_probabilities=dict(zip(categories.tolist(), probabilities.tolist())),
        car_count_random_seed=random_seed,
    )
    return result
